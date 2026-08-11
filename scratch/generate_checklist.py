import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Reconfigurar stdout para utf-8
sys.stdout.reconfigure(encoding='utf-8')

# Criar workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Checklist Homologação"

# Habilitar grid lines
ws.views.sheetView[0].showGridLines = True

# Paleta de Cores Premium (Azul Escuro / Ciano / Cinza)
fill_header = PatternFill(start_color="0A1128", end_color="0A1128", fill_type="solid")
fill_sub_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

font_title = Font(name="Segoe UI", size=16, bold=True, color="00D4FF")
font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
font_sub_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Segoe UI", size=10, bold=True, color="000000")
font_regular = Font(name="Segoe UI", size=10, color="000000")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color="E2E8F0"),
    right=Side(style='thin', color="E2E8F0"),
    top=Side(style='thin', color="E2E8F0"),
    bottom=Side(style='thin', color="E2E8F0")
)

# Título Principal
ws.merge_cells("A1:G1")
title_cell = ws["A1"]
title_cell.value = "📋 CHECKLIST DE TESTE E HOMOLOGAÇÃO MANUAL - CR$ HOME CARE AI"
title_cell.font = font_title
title_cell.fill = fill_header
title_cell.alignment = align_center
ws.row_dimensions[1].height = 40

# Cabeçalhos das Colunas
headers = [
    "ID",
    "Módulo / Tela",
    "Caso de Teste / Funcionalidade",
    "Passos para Reproduzir",
    "Resultado Esperado",
    "Status (Aprovado/Reprovado/Pendente)",
    "Observações / Detalhes de Bugs"
]

for col_idx, text in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_idx)
    cell.value = text
    cell.font = font_header
    cell.fill = fill_sub_header
    cell.alignment = align_center
    cell.border = thin_border
ws.row_dimensions[2].height = 25

# Dados do Checklist
checklist_data = [
    # Módulo: Autenticação e Geral
    ("CT-001", "Acesso & Geral", "Acesso à página inicial (Landing Page)", 
     "1. Acessar a URL base da aplicação (http://localhost:8000/ ou /landing).\n2. Verificar se carrega a Landing Page corretamente com depoimentos, planos e links de termos.",
     "Página é carregada rapidamente, sem erros de console e com design premium responsivo.", "", ""),
    
    ("CT-002", "Autenticação", "Login de Cliente (Paciente)",
     "1. Clicar em Entrar / Área do Cliente.\n2. Inserir email e senha de um paciente cadastrado.\n3. Clicar em Entrar.",
     "Redireciona com sucesso para '/dashboard-cliente' carregando os dados do paciente.", "", ""),

    ("CT-003", "Autenticação", "Login de Administrador",
     "1. Acessar a rota de login ou área do administrador.\n2. Entrar com dados de admin master.\n3. Clicar em Entrar.",
     "Redireciona com sucesso para '/dashboard' (painel de administração).", "", ""),

    ("CT-004", "Autenticação", "Recuperação / Fluxo de Senha Incorreta",
     "1. Tentar fazer login com senha inválida.\n2. Tentar login com email não cadastrado.",
     "Sistema exibe alerta/notificação amigável de erro e não permite acesso.", "", ""),

    ("CT-005", "Autenticação", "Logout (Sair)",
     "1. No cabeçalho de qualquer dashboard, clicar no botão vermelho 'Sair'.",
     "Limpa o localStorage, encerra a sessão e redireciona de volta para a tela de login inicial.", "", ""),

    # Módulo: Dashboard Cliente
    ("CT-006", "Dashboard Cliente", "Visualização da Agenda do Dia",
     "1. Visualizar o cabeçalho principal e agenda do cliente.\n2. Verificar se a data exibida no topo ('Minha agenda para hoje dia: DD/mês/AAAA') condiz com a data atual selecionada.\n3. Conferir os contadores de compromissos.",
     "A data exibida e os contadores de medicamentos/consultas batem exatamente com as informações salvas no banco para aquele dia específico.", "", ""),

    ("CT-007", "Dashboard Cliente", "Visualização Responsiva de Medicamentos",
     "1. Abrir a página do cliente em resoluções de desktop, tablet e celular (ou simular no Chrome DevTools).\n2. Focar na seção de medicamentos ('Remédios para hoje').",
     "Nenhum card ou texto é cortado à esquerda. Em telas móveis, os cards e botões são empilhados responsivamente respeitando as margens laterais.", "", ""),

    ("CT-008", "Dashboard Cliente", "Adicionar Medicamento (Formulário Comum)",
     "1. Clicar no botão '➕ Novo' na seção de remédios.\n2. Preencher nome, dosagem, horário (hora/minuto) e duração em dias.\n3. Selecionar dias da semana.\n4. Clicar em Salvar.",
     "Medicamento é salvo e aparece imediatamente na lista com o horário e detalhes configurados.", "", ""),

    ("CT-009", "Dashboard Cliente", "Fotografar Receita (OCR com IA)",
     "1. Clicar em '➕ Novo' e depois em 'Fotografar Receita Médica'.\n2. Fazer upload de uma imagem de receita médica.\n3. Aguardar a IA analisar.",
     "A IA processa a receita, identifica os medicamentos e pré-preenche os campos do formulário automaticamente.", "", ""),

    ("CT-010", "Dashboard Cliente", "Histórico e Independência de Horários por Data",
     "1. Mudar o dia no calendário.\n2. Confirmar um medicamento como 'Tomei' no dia de hoje.\n3. Navegar para outro dia e verificar o status do mesmo medicamento.",
     "A edição/confirmação em um dia NÃO afeta o status ou o agendamento de outros dias (independência real de datas no banco).", "", ""),

    ("CT-011", "Dashboard Cliente", "Uso Contínuo e Alerta de Revisão",
     "1. Cadastrar um medicamento como 'Uso Contínuo' com duração definida (ex: 6 meses).\n2. Simular/verificar após a data limite.",
     "O sistema exibe o alerta amarelo pulsante de revisão de receita médica para medicamentos contínuos vencidos.", "", ""),

    ("CT-012", "Dashboard Cliente", "Anexar Foto da Caixa do Remédio",
     "1. Clicar no ícone de câmera (📷) em um card de medicamento.\n2. Fazer o upload de uma imagem da caixa.\n3. Verificar se a miniatura aparece no card.",
     "A miniatura da imagem é salva no banco, exibida no card do cliente e pode ser clicada para visualização expandida.", "", ""),

    ("CT-013", "Dashboard Cliente", "Confirmar Medicamento Tomado/Não Tomado",
     "1. Clicar no botão 'Tomei' em um medicamento pendente.\n2. Clicar no botão 'Não Tomei' (ou simular atraso).\n3. Testar o botão 'Reagendar'.",
     "Se tomado, o card fica verde com check. Se não tomado, abre modal permitindo reagendar horário específico ou marcar como definitivo não tomado (card vermelho).", "", ""),

    ("CT-014", "Dashboard Cliente", "Gerenciamento de Consultas",
     "1. Na seção de Consultas, clicar em '➕ Novo'.\n2. Preencher especialidade, médico, data, hora, local e observações.\n3. Salvar e depois tentar editar e excluir.",
     "A consulta aparece listada, exibe o local/observações, e permite edição/exclusão em tempo real.", "", ""),

    ("CT-015", "Dashboard Cliente", "Chat IA - Maximus (Texto e Voz)",
     "1. Digitar uma dúvida no input do Maximus e clicar em Enviar.\n2. Clicar em uma sugestão rápida.\n3. Ativar 'Falar respostas' e enviar mensagem.\n4. Testar o botão de microfone para entrada de voz.",
     "O Maximus responde via chat contextualizado, gera áudio da resposta (lido em voz alta se habilitado) e aceita comandos de voz.", "", ""),

    # Módulo: Painel Admin
    ("CT-016", "Dashboard Admin", "Listagem e Detalhes dos Clientes",
     "1. Acessar o Dashboard Admin.\n2. Verificar a lista de clientes cadastrados, contadores e o status de agendamentos.\n3. Clicar em um cliente da lista.",
     "Mostra a lista completa com miniestatísticas precisas para cada cliente (remédios, consultas, responsáveis).", "", ""),

    ("CT-017", "Dashboard Admin", "Gerenciamento de Assinaturas e Planos",
     "1. Acessar o painel de planos ou simular a criação de uma assinatura via Mercado Pago.",
     "Integração do checkout de assinatura exibe o QR Code Pix/cartão e atualiza o status do cliente após confirmação.", "", ""),

    # Módulo: Notificações & Sistema
    ("CT-018", "Notificações", "Envio de Alertas via WhatsApp",
     "1. Cadastrar medicamento para o horário atual.\n2. Aguardar o scheduler do servidor rodar a verificação de minuto a minuto.",
     "O servidor identifica o medicamento pendente e dispara notificação automática no WhatsApp cadastrado do usuário.", "", ""),

    ("CT-019", "Sistema", "Visualização Nativa de PDFs",
     "1. Tentar visualizar um documento anexado do paciente (carteirinha, exames).\n2. Clicar no link de visualização.",
     "O PDF abre diretamente e de forma nativa na aba do navegador via rota HTTP inline (sem forçar download direto).", "", ""),

    ("CT-020", "Sistema", "Fuso Horário Estrito de Brasília",
     "1. Cadastrar medicamentos e consultas.\n2. Verificar os horários registrados no banco de dados e as notificações.",
     "Todas as datas e horários seguem estritamente o fuso horário de Brasília (UTC-3), impedindo inconsistências entre servidor e cliente.", "", ""),
]

# Escrever os dados e aplicar formatação
current_row = 3
for row_data in checklist_data:
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = value
        cell.border = thin_border
        
        # Alinhamento
        if col_idx in [1, 2, 6]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left
            
        # Cores e Fontes
        cell.font = font_regular
        if current_row % 2 == 0:
            cell.fill = fill_zebra
            
    ws.row_dimensions[current_row].height = 55
    current_row += 1

# Ajustar largura das colunas
column_widths = {
    "A": 10,  # ID
    "B": 20,  # Módulo
    "C": 35,  # Caso de Teste
    "D": 50,  # Passos
    "E": 45,  # Resultado Esperado
    "F": 25,  # Status
    "G": 30   # Observações
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Salvar planilha
output_path = "checklist_homologacao.xlsx"
wb.save(output_path)
print(f"✅ Planilha '{output_path}' gerada com sucesso!")
